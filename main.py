from sqlalchemy.engine.url import make_url
import MySQLdb
import os
from openpyxl import Workbook
from datetime import date, datetime

REQUEST_RESOURCE = 1
RR_MAP = ["REQUEST", "RESOURCE"]


url = make_url(os.environ.get("CLEARDB_DATABASE_URL")) # parse database url into separate parts like port, database, username ...

con = MySQLdb.Connection(           # get connection to db, analogous to php mysqli()
    host=url.host,
    user=url.username,
    passwd=url.password,
    port=3306,
    db=url.database
    )

con.query("""
    SELECT 
        p.post_id,
        p.request_resource,
        p.state,
        p.city,
        p.description,
        p.time,
        p.upvotes,
        p.downvotes,
        p.verified,
        p.email,
        p.ph_no,
        group_concat(t.tag_name) as needs,
        p.first_name,
        p.last_name
    FROM 
        post as p,
        needs as n,
        tag as t
    WHERE
        p.post_id=n.post_id AND
        n.tag_id=t.tag_id
    GROUP BY
        p.post_id
    """)    
     
posts = con.store_result()         # store results of query into result
workbook = Workbook()               #create an excel workbook instance
sheet = workbook.active                #get current active sheet in excel workbook
column_names = [x[0] for x in posts.describe()]    #get all collumn names from result

for col_index,col_name in enumerate(column_names):  
    sheet.cell(row=1, column=col_index+1, value=col_name )    #write table column names into the excel file

row=2
col=1

for _ in range(posts.num_rows()):
    post = posts.fetch_row()
    post = list(post[0])
    post[REQUEST_RESOURCE] = RR_MAP[int(post[REQUEST_RESOURCE])]
    for val in post:
        sheet.cell(row=row,column=col,value=str(val))
        col = col + 1
    row = row + 1
    col = 1


datetime_string = datetime.utcnow().strftime("%Y-%m-%d_%H-%M-%S")
filename = datetime_string + ".xlsx"
workbook.save(filename=filename)